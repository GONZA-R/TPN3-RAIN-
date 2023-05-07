


############################
#Borrar pantalla
import os
def clear_screen():
    os.system('clear' if os.name == 'posix' else 'cls')
############################

#####################################################################################################
# Funciones punto 1

import urllib
from bs4 import BeautifulSoup


def obtener_enlaces(url):
    html_pagina = urllib.request.urlopen(url)#devuelve la pagina formato html
    soup = BeautifulSoup(html_pagina, features="html.parser")
    etiquetas = soup('a') #en html los link estan referenciado por una etiqueta a
    #etiquetas es una tipo lista que devuelve 

    ### Aqui trabaja sobre cada una de las urls obtenidas
    lista_de_urls=[]
    for tag in etiquetas:
        url_html=tag.get('href') #devuelve una cadena de texto
        try:
            if url_html[0:4] == 'http':
                url_completa=str(tag.get('href'))#cadena de texto
            else :
                url_completa=(url.rstrip("/")+str(tag.get('href')))#concatena  la direccion relativa con la direccion original
            lista_de_urls.append(url_completa)
        except:
            print('')
    return lista_de_urls


import urllib.request

def obtener_enlaces_secundarios(lista_de_urls):
    i=0
    dic_de_url = {}
    for url in lista_de_urls:
        lista_urls_secundarias = []
        i=i+1
        print(f"Accediendo a los enlaces dentro de la página N° {str(i)}   {url}")
        try:     
            html_pagina = urllib.request.urlopen(url)
            soup = BeautifulSoup(html_pagina, features="html.parser")
            etiquetas_secu = soup("a")  # devuelve todas las urls con etiqueta 'a' que están dentro de la página
            
            if len(etiquetas_secu) > 0: 
                #print(f"El link de 1er Nivel posee una cantidad de {len(etiquetas_secu)} Links de 2do Nivel\n")
                for nueva_eti in etiquetas_secu:
                    url_html2 = nueva_eti.get("href")
                    try:
                        if url_html2.startswith("http"):
                            url_completa_secu = str(nueva_eti.get("href"))
                        else:
                            url_completa_secu = url + str(nueva_eti.get("href"))
                        lista_urls_secundarias.append(url_completa_secu)
                    except:
                        print("")

                lista_urls_secundarias = list(set(lista_urls_secundarias))
                lista_urls_secundarias = sorted(lista_urls_secundarias)
                dic_de_url[url] = lista_urls_secundarias
            else:
                print("No hay más enlaces")
        except:
                print("Algo ha fallado")  # si la petición al servidor falla, salta al except
            
    return dic_de_url
#####################################################################################################
import pandas as pd
        
        
def exportar_diccionario_a_excel(diccionario, nombre_archivo):
        # Crear un DataFrame a partir del diccionario
        df = pd.DataFrame.from_dict(diccionario, orient='index')

        # Transponer el DataFrame para que las claves sean filas y los valores sean columnas
        df = df.transpose()

        # Exportar el DataFrame a un archivo de Excel
        df.to_excel(nombre_archivo, index=False)


#####################################################################################################
from openpyxl import load_workbook

def modificar_formato_columnas_xlsx(nom_archivo):
    # Cargar el archivo existente
    book = load_workbook(nom_archivo)

    # Seleccionar la hoja a modificar
    sheet = book.active

    # Modificar el formato de las columnas
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # Guardar los cambios en el archivo existente
    book.save(nom_archivo)

#####################################################################################################
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

def colorear_primer_fila_excel(archivo_excel):
    """
    Función que colorea la primera fila de un archivo Excel
    """
    # Cargar el archivo Excel
    wb = load_workbook(archivo_excel)
    # Seleccionar la hoja de trabajo
    ws = wb.active
    # Definir el patrón de relleno
    fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    # Obtener la primera fila
    row = ws[1]
    # Aplicar el patrón de relleno a cada celda de la primera fila
    for cell in row:
        cell.fill = fill
    # Guardar el archivo Excel
    wb.save(archivo_excel)
# Finaliza funciones punto 1
#####################################################################################################



#####################################################################################################
# Funciones punto 2

import requests
from bs4 import BeautifulSoup
def conseguir_url(url):
            response = requests.get(url)
            soup = BeautifulSoup(response.text, 'html.parser')
            urls_noticias=[]
            divs = soup.find_all('div', {'class': 'd23_content-section'}) #Segun la estructura html de la pagina
            #los links a necesitar se encuentran en la class = d23_content-section
            for div in divs:
                links = div.find_all('a')
                for link in links:
                    href = link.get('href')
                    if href and 'autor' not in href and 'infobae' not in href:
                        if href and 'economia/2023' in href:
                            #print(href)
                            urls_noticias.append(href)
            return urls_noticias
#####################################################################################################

import requests
from bs4 import BeautifulSoup

def web_scrapping(links):
    noticias = []
    for link in links:
            
        url = link
        response = requests.get(url)
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
            
        # Obtener título de la noticia
        titulo = soup.find('h1').text.strip()

        #obtener subtitulo de la noticia
        resumen = soup.find('h2').text.strip()
            
        # Obtener contenido de la noticia
        parrafos = [p.text for p in soup.find_all("p")]


        img_principales = soup.find('div', {'class': 'd23-body-article'})
        if img_principales:
            img_principales = img_principales.find_all('img')
        else:
            img_principales = None
        url_imagen_principal = [img['src'] for img in img_principales] if img_principales else []
        
        # Diccionario con cada elemento de la pagina a consultar
        noticias.append({'titulo': titulo,'resumen': resumen, 'contenido': parrafos,'url_imagenes':url_imagen_principal}) 
    i=0
        # guardar archivo de texto de las 10 primeras noticias
    for noticia in noticias:
        i=i+1
        nombre_noticia='Noticia N° '+str(i)+'.txt'
        guardar_noticias(nombre_noticia,noticia)
        if i>=3:#original i>=10 para 10 noticias cambiar despues al original#####################333
            break
        else:
            pass
    return noticias

#####################################################################################################

def guardar_noticias(nombre_archivo,noticia):
    with open(nombre_archivo, 'w') as file:
        file.write('Título: \n' + noticia['titulo'] + '\n\n')
        file.write('Resumen: \n' + noticia['resumen'] + '\n\n')
        file.write('Contenido: \n\n')
        for parrafo in noticia['contenido']:
            file.write(parrafo + '\n')
        file.write('\nURLs de las imagenes:\n\n')
        for url in noticia['url_imagenes']:
            file.write(url + '\n')
        file.write('\n')
#######################################################################

import string
from nltk.corpus import stopwords

def eliminar_puntuaciones(tokens):
    # Obtener los signos de puntuación
    signos_puntuacion = set(string.punctuation)
    # Crear una nueva lista sin signos de puntuación
    tokens_sin_puntuacion = [token for token in tokens if token not in signos_puntuacion]
    return tokens_sin_puntuacion

####################################################################
def eliminar_stopwords(lista_palabras):
    # Obtenemos las stopwords para español
    stop_words = set(stopwords.words('spanish'))
    
    # Eliminamos las stopwords de la lista de palabras
    lista_sin_stopwords = [palabra for palabra in lista_palabras if not palabra.lower() in stop_words]
    
    return lista_sin_stopwords
#####################################################################
import re
def eliminar_links(texto):
    # Expresión regular para buscar URLs
    patron = re.compile(r'https?://\S+')
    # Remover URLs del texto
    texto_sin_urls = patron.sub('', texto)
    return texto_sin_urls
#####################################################################
import re

def corregir_barras(texto):
    # Buscar patrones de barra y reemplazarlos con un espacio
    texto = re.sub(r'/', ' ', texto)
    
    return texto

import re

def corregir_guion(texto):
    # Buscar patrones de guion seguido o precedido por espacios y eliminarlos
    texto = re.sub(r'\s?-\s?', ' ', texto)
    
    # Buscar patrones de guion seguido o precedido por caracteres no alfanuméricos y eliminarlos
    texto = re.sub(r'(\W+)-\s?|(\s?)-(\W+)', r'\1\2\3', texto)
    
    # Buscar patrones de guion al final o al inicio de palabras y eliminarlos
    texto = re.sub(r'(\b\w+)-\b|\b-(\w+\b)', r'\1\2', texto)
    
    return texto

#####################################################################


from collections import Counter
def obtener_frecuencias(palabras):
    # Contar la frecuencia de cada palabra
    frecuencias = Counter(palabras)
    # Devolver la lista de términos más frecuentes y sus frecuencias
    return frecuencias.most_common(100)

######################################################################
from nltk.stem import SnowballStemmer

def algoritmo_snowball(lista_palabras):
    stemmer = SnowballStemmer('spanish')
    stemmed = [stemmer.stem(palabra) for palabra in lista_palabras]
    return stemmed
#######################################################################
def escribir_frecuencias_en_txt(frecuencias, nombre_archivo):
    with open(nombre_archivo, 'w') as archivo:
        for lista_frecuencias in frecuencias:
            linea = '\t'.join([str(palabra) + ':' + str(frecuencia) for palabra, frecuencia in lista_frecuencias]) + '\n'
            archivo.write(linea)
            archivo.write('\n\n')

#######################################################################
def escribir_lista_en_txt(lista, nombre_archivo):
    with open(nombre_archivo, 'w') as archivo:
        for elemento in lista:
            archivo.write(str(elemento) + '\n')




# Finaliza funciones punto 2
#####################################################################################################

#####################################################################################################
# Funciones punto 3

import nltk
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity


def compare_news(news1, news2):
    # Convertir las listas tokenizadas a strings
    text1 = " ".join(news1)
    text2 = " ".join(news2)

    # Lemmatización de los tokens
    lemmatizer = WordNetLemmatizer()
    tokens1 = word_tokenize(text1)
    tokens2 = word_tokenize(text2)
    tokens1 = [lemmatizer.lemmatize(word) for word in tokens1]
    tokens2 = [lemmatizer.lemmatize(word) for word in tokens2]

    # Verificar que las listas de tokens no estén vacías
    if not tokens1 or not tokens2:
        return 0.0

    # Vectorización de los textos utilizando la medida TF-IDF
    vectorizer = TfidfVectorizer()
    tfidf1 = vectorizer.fit_transform([text1])
    tfidf2 = vectorizer.transform([text2])

    # Cálculo de la similitud coseno entre los vectores TF-IDF de los textos
    sim_cos = cosine_similarity(tfidf1, tfidf2)

    return sim_cos[0][0]
#####################################################################################################

import openpyxl

def crear_matriz_similitud_excel(titulos,palabras_noticias):
    # Obtener el número de noticias
    num_noticias = len(titulos)


    # Crear un libro y seleccionar la hoja
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Similitud"

    # Asignar nombres a las filas y columnas
    for i in range(num_noticias):
        hoja.cell(row=1, column=i+2, value=f"Noticia {i+1}")
        hoja.cell(row=i+2, column=1, value=f"Noticia {i+1}")

    # Llenar la diagonal principal con 1
    for i in range(num_noticias):
        hoja.cell(row=i+2, column=i+2, value=1)

    # Calcular las similitudes y escribirlas en la matriz
    for i in range(num_noticias):
        for j in range(i+1, num_noticias):
            similitud = compare_news(palabras_noticias[i], palabras_noticias[j])
            hoja.cell(row=i+2, column=j+2, value=round(similitud, 2))
            hoja.cell(row=j+2, column=i+2, value=round(similitud, 2))

    # Guardar el archivo
    libro.save("matriz_similitud.xlsx")




#####################################################################################################
import os
import openpyxl





# Main 
while True:
    clear_screen()
    print("Trabajo Practico N°3 Crawler y Scraper"+"\n\n")
    print("1. Punto 1: Desarrollo de Crawler para recolectar URLs de primer y segundo nivel de un sitio WEB")
    print("2. Punto 2: Web Scraping y Análisis Textual de las 10 Primeras Noticias de un sitio WEB")
    print("3. Punto 3: Identificación de Similitudes y Relaciones entre Noticias a través de Términos en Común")
    print("4. Salir\n")

    opcion = input("Ingrese una opción: ")
    if opcion == "1":
        """
        Planifique, diseñe y construya un crawler para recolectar todas las URLs de los primeros 
        2 niveles de profundidad del sitio web:https://www.fi.unju.edu.ar/
        Deberá crear una hoja de cálculo (Excel o Google SpreadSheet) para almacenar lo recolectado, 
        teniendo en cuenta de identificar las URLs de cada nivel de origen. Evite recolectar URLs repetidas, 
        para ello deberá almacenar de algún modo las URLs que vaya visitando.

        """
        clear_screen()

        url = "https://www.fi.unju.edu.ar/" #Defino la url del link a visitar
        lista_de_urls=obtener_enlaces(url) #Aqui obtiene todas las urls de primer nivel

        lista_de_urls = list(set(lista_de_urls)) # eliminar urls repetidas
        lista_de_urls = sorted(lista_de_urls) # ordenarlas
        i=0
        print('Enlaces de página principal: \r\n')
        for tag in lista_de_urls:
            i=i+1
            print("Pagina N° "+str(i)+" "+tag) 
        print('Tamaño de lista:', len(lista_de_urls))

        print('\nEnlaces de 2do NIVEL: \r\n')
        diccionario_de_urls = obtener_enlaces_secundarios(lista_de_urls)#Apartir de aqui se obtiene un diccionario
        #Que contiene como clave las url y como valores la listas de urls securndarias para cada link
        
###########################################################################
        # Exportamos el diccionario como archivo EXCEL
        nombre_archivo="Archivo_Excel_Punto1.xlsx"
        exportar_diccionario_a_excel(diccionario_de_urls,nombre_archivo)

##########################################################################3
        #Necesario para dar un formato de visualizacion al archivo excel
        modificar_formato_columnas_xlsx(nombre_archivo)

############################################################################
        #Dar color a la primer fila del excel
        colorear_primer_fila_excel(nombre_archivo)
############################################################################


        input("Presione enter para continuar...")
        pass

    elif opcion == "2":

        """
        Realice un web scraping de la siguiente URL:
        https://www.infobae.com/economia/
        De esta URL recolecte las primeras 10 noticias, identificando por cada una el Título, 
        Resumen, Listado de imágenes (ubicación del archivo) y el Cuerpo de la misma. 
        A continuación realice un análisis textual sencillo, tokenize dichos documentos, 
        elimine las stop-words y liste los 100 términos más frecuentes. 
        En el mismo sentido realice un stemming y vuelva a listar los 100 términos más frecuentes. 
        """

        clear_screen()#Borra pantalla

        url = 'https://www.infobae.com/economia/'
        lista_de_noticias=conseguir_url(url)
        lista_de_noticias = list(set(lista_de_noticias))
        lista_de_noticias.sort()

        
        url_base = 'https://www.infobae.com'
        lista_url_completa=[] #la url completa de las noticias va estar formado por el url base + cada link de
        #la lista de noticias

        print('Accediendo a las siguientes paginas...\n')
        for noticia in lista_de_noticias:
            print(url_base+noticia+'\n')
            lista_url_completa.append(url_base+noticia)
        #######################################################################

        #print(lista_url_completa)
        dic_noticias=web_scrapping(lista_url_completa)#Aqui se llama a la funcion que se encarga de traer los titulos,resumenes
        #contenido de los parrafos y lista de imagenes para guardar todo en un documento de texto
        with open('Lista de URLs.txt', 'w') as file:#Va a guardar la lista en un archivo de texto para 
            #visualizar mejor con que links se va a trabajar
            file.write('\n'.join(lista_url_completa))
        print("\nSe genero un archivo de texto...\n")


        #Apartir de aqui ya esta los txt con las noticias, ahora falta procesarlos para seguir trabajando
        import nltk
        from nltk.corpus import stopwords
        from nltk.tokenize import word_tokenize
        
        j=0
        palabras_noticias=[]
        palabras_noticias_stemming=[]
        lista_palabras_noticias=[]
        lista_palabras_stemming=[]

        palabras_noticias_aux=[]

        for i in range(1,4):#original (1,11) para 10 noticias cambiar despues al original
            # aquí va el código que se ejecutará en cada iteración
            j=j+1
            nombre="Noticia N° "+str(j)+".txt"
            with open(nombre, 'r') as file:
                texto = file.read()#abre el archivo y va generando una cadena de texto
                texto = eliminar_links(texto)#elimino los link que existen en la cadena de texto
                texto= texto.lower()#convierte toda la cadena de texto en minuscula para poder trabajar bien
                texto= re.sub(r'\d+\.?\d*', '', texto)#elimina los numeros enteros y en decimales
                texto=corregir_guion(texto)
                texto=corregir_barras(texto)
                texto = texto.replace('“', '')#elimina comillas dobles que quedaron en el texto
                texto = texto.replace('”', '')#elimina comillas dobles que quedaron en el texto
                
                

            #Esta parte empezamos a procesar procesar el texto para trabajarlo
            
            palabras_noticias=word_tokenize(texto)
            palabras_noticias=eliminar_puntuaciones(palabras_noticias)
            palabras_noticias=eliminar_stopwords(palabras_noticias)


            palabras_noticias_stemming=palabras_noticias
            

            ################################################################################################
            escribir_lista_en_txt(palabras_noticias, "Token Noticia N° "+str(j)+".txt")
            palabras_noticias_aux.append(palabras_noticias)###############################sacar si esta mal
            #################################################################################################

            palabras_noticias=obtener_frecuencias(palabras_noticias)

            palabras_noticias_stemming=algoritmo_snowball(palabras_noticias_stemming)
            palabras_noticias_stemming=obtener_frecuencias(palabras_noticias_stemming)

            
            lista_palabras_noticias.append(palabras_noticias)
            lista_palabras_stemming.append(palabras_noticias_stemming)

        """
        #continuar aqui
        print('\nLista de los 100 terminos mas frecuentes de las palabras de cada una de las 10 noticias')
        for elemento in lista_palabras_noticias:
            print('\n')
            print(elemento)
        
        escribir_frecuencias_en_txt(lista_palabras_noticias, 'Frecuencias de las palabras.txt')
        print('\nLista de los 100 terminos mas frecuentes de cada raiz de las palabras de cada una de las 10 noticia')
        for elemento in lista_palabras_stemming:
            print('\n')
            print(elemento)
        escribir_frecuencias_en_txt(lista_palabras_stemming, 'Frecuencias de las raices.txt')
        """#esto despues sacar de los comentarios, solo es para acelerar el proceso de trabajo

############################################################################
        

        input("Presione enter para continuar...")

        pass


    elif opcion == "3":

        """Sería capaz de identificar si alguna de las noticias es muy parecida a otra o 
        están muy relacionadas por la existencia o co-existencia de términos en común? 
        """
    
        clear_screen()
        
        titulos = range(len(palabras_noticias_aux))
        crear_matriz_similitud_excel(titulos,palabras_noticias_aux)
        max_similitud = 0.0

        for i in range(len(titulos)):
            for j in range(i+1, len(titulos)):
                similitud = compare_news(palabras_noticias_aux[i], palabras_noticias_aux[j])
                sim1=similitud
                max_similitud = max(max_similitud, sim1)
                print("Similitud coseno: {:.2f} entre la noticia {} con la noticia {} asi tambien entre la noticia {} con la noticia {}".format(sim1, i+1, j+1, j+1, i+1))
            

        print("\nLa mayor similitud es {:.2f}".format(max_similitud))
        print("\nSe genero un archivo excel para estudiar mejor los resultados...")
        


    
    ###################################################################################################
        
    ###################################################################################################



####################################################################################3
        


        input("Presione enter para continuar...")

        pass

    elif opcion == "4":
        clear_screen()
        print("Saliendo del programa...")
        break
    else:
        print("Opción no válida, por favor intente de nuevo.")
        input("Presione enter para continuar...")
